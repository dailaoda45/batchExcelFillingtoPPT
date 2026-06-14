# -*- coding: utf-8 -*-
"""
逐行读取 data.xlsx，每行生成一份独立的 PPTX 文件。

工作流程：
  1. 打开 模版.pptx，读取第一页的形状信息（文本框尺寸、字体大小等）
  2. 读取 data.xlsx，按行获取数据（标题、副标题）
  3. 对每一行数据：
     a. 复制模版全部文件
     b. 替换幻灯片中的占位符文字
     c. 根据文本框尺寸和文字长度，自动计算合适的字体大小（模拟 AutoFit）
     d. 设置段落居中对齐
     e. 保存为 {文件夹名称}.pptx

模版中的占位符：
  【文档关键字[标题]】  →  替换为 data.xlsx 的"标题"列
  【文档关键字[副标题]】 →  替换为 data.xlsx 的"副标题"列
"""

import os
import copy
import zipfile
import atexit

from lxml import etree
from fontTools.ttLib import TTCollection
from PIL import ImageFont
import openpyxl


# ═══════════════════════════════════════════════════════════════
# 用户配置（修改这里来控制行为）
# ═══════════════════════════════════════════════════════════════

MAX_ROWS = 0  # ★ 测试时改成 5（只处理前5行），正式跑改成 0（全部）


# ═══════════════════════════════════════════════════════════════
# XML 命名空间常量
# （PPTX 本质是 ZIP 包，内部用 XML 描述内容，这些是 XML 的命名空间）
# ═══════════════════════════════════════════════════════════════

P_SP   = '{http://schemas.openxmlformats.org/presentationml/2006/main}sp'     # 幻灯片中的形状
A_R    = '{http://schemas.openxmlformats.org/drawingml/2006/main}r'           # 文本 run（一段连续格式的文字）
A_T    = '{http://schemas.openxmlformats.org/drawingml/2006/main}t'           # 文本内容
A_RPR  = '{http://schemas.openxmlformats.org/drawingml/2006/main}rPr'         # run 的属性（字体、大小、颜色等）
A_XFRM = '{http://schemas.openxmlformats.org/drawingml/2006/main}xfrm'        # 形状的变换（位置、大小）
A_EXT  = '{http://schemas.openxmlformats.org/drawingml/2006/main}ext'         # 尺寸（宽度、高度）
A_P    = '{http://schemas.openxmlformats.org/drawingml/2006/main}p'           # 段落
A_PPR  = '{http://schemas.openxmlformats.org/drawingml/2006/main}pPr'         # 段落属性（对齐方式等）


# ═══════════════════════════════════════════════════════════════
# 字体加载
# 微软雅黑是 .ttc（TrueType Collection）格式，不能直接用 Pillow 读取，
# 需要用 fontTools 把里面的单个字体提取为 .ttf 再使用。
# ═══════════════════════════════════════════════════════════════

FONT_TTC    = 'C:\\Windows\\Fonts\\msyh.ttc'     # 微软雅黑（常规体）
FONT_TTC_BD = 'C:\\Windows\\Fonts\\msyhbd.ttc'   # 微软雅黑（粗体）
TMP_FONTS   = '_tmp_fonts'                        # 临时存放提取出的 .ttf 文件

os.makedirs(TMP_FONTS, exist_ok=True)

def _extract_ttf(ttc_path, index=0):
    """从 TTC 文件中提取指定索引的 TTF 并缓存到临时目录"""
    name = os.path.splitext(os.path.basename(ttc_path))[0]
    out = os.path.join(TMP_FONTS, f'{name}_{index}.ttf')
    if not os.path.exists(out):
        ttc = TTCollection(ttc_path)
        font = ttc.fonts[index]
        # 删除 MERG 表（Apple 遗留格式，会影响 fontTools 的子集化）
        if 'MERG' in font:
            del font['MERG']
        font.save(out)
    return out

FONT_TTF    = _extract_ttf(FONT_TTC)             # → 微软雅黑常规体 .ttf
FONT_TTF_BD = _extract_ttf(FONT_TTC_BD)          # → 微软雅黑粗体 .ttf

# 程序退出时自动清理临时字体文件
def _cleanup_fonts():
    if os.path.exists(TMP_FONTS):
        for f in os.listdir(TMP_FONTS):
            os.remove(os.path.join(TMP_FONTS, f))
        os.rmdir(TMP_FONTS)
atexit.register(_cleanup_fonts)


# ═══════════════════════════════════════════════════════════════
# 字体大小计算（代码实现 AutoFit）
# ═══════════════════════════════════════════════════════════════

def measure_text_width(text, font_path, size_pt):
    """
    用 Pillow 测量文本在指定字号下的像素宽度。
    在 72 DPI 下，1pt = 1px，所以返回值可以直接当作磅值使用。
    """
    font = ImageFont.truetype(font_path, int(size_pt))
    bbox = font.getbbox(text)
    return bbox[2] - bbox[0]


def calc_multiline(text, font_path, box_w_pt, box_h_pt, start_sz, line_spacing=1.3):
    """
    计算适合文本框的字体大小（支持多行自动换行）。

    参数：
      text        — 要显示的文本
      font_path   — 字体文件路径
      box_w_pt    — 文本框宽度（磅）
      box_h_pt    — 文本框高度（磅）
      start_sz    — 起始字号（从多大开始向下尝试）
      line_spacing — 行距系数，1.3 表示行高 = 字号 × 1.3

    返回值：
      (最佳字号, 行数)

    算法：
      从 start_sz 开始向下尝试每个字号：
        1. 用 Pillow 逐字测量宽度，模拟文本自动换行
        2. 计算所需的行数
        3. 计算总高度 = 行数 × 行高
        4. 如果总高度 ≤ 文本框高度，则当前字号合适
    """
    if not text:
        return start_sz, 0

    for sz in range(int(start_sz), 7, -1):
        font = ImageFont.truetype(font_path, sz)
        line_height = sz * line_spacing

        # 模拟自动换行 —— 逐字累加宽度，超宽则换行
        lines = []
        current_line = ''
        current_width = 0
        for ch in text:
            char_width = font.getbbox(ch)[2]
            # 留 5% 边距，避免文字贴边
            if current_width + char_width > box_w_pt * 0.95 and current_line:
                lines.append(current_line)
                current_line = ch
                current_width = char_width
            else:
                current_line += ch
                current_width += char_width
        if current_line:
            lines.append(current_line)

        total_height = len(lines) * line_height
        if total_height <= box_h_pt:
            return sz, len(lines)

    return 7, 99  # 极端情况：最小 7pt


# ═══════════════════════════════════════════════════════════════
# 第一步：读取 data.xlsx
# ═══════════════════════════════════════════════════════════════

wb = openpyxl.load_workbook('data.xlsx', data_only=True)
ws = wb.active

rows = []                                   # [(文件夹名称, 标题, 副标题), ...]
row_limit = MAX_ROWS if MAX_ROWS > 0 else ws.max_row - 1

for r in range(2, 2 + row_limit):
    if r > ws.max_row:
        break
    folder   = str(ws.cell(r, 1).value or '').strip()
    title    = str(ws.cell(r, 2).value or '').strip()
    subtitle = str(ws.cell(r, 3).value or '').strip()
    if folder:
        rows.append((folder, title, subtitle))

print(f"共读取 {len(rows)} 条数据")


# ═══════════════════════════════════════════════════════════════
# 第二步：读取模版，分析第一页的形状信息
# ═══════════════════════════════════════════════════════════════

with zipfile.ZipFile('模版.pptx', 'r') as z:
    # 将模版的所有文件读入内存字典（文件名 → 字节数据）
    template_files = {name: z.read(name) for name in z.namelist()}

# 解析第一张幻灯片的 XML
slide1_xml = etree.fromstring(template_files['ppt/slides/slide1.xml'])

# 遍历幻灯片中的所有形状，识别出"标题"和"副标题"两个文本框
shape_infos = []     # 存储每个有效形状的元信息
for sp in slide1_xml.iter(P_SP):
    # 收集该形状内所有文本，合并后判断角色
    t_elems = list(sp.iter(A_T))
    combined = ''.join(t.text or '' for t in t_elems)

    if '标题' in combined and '副标题' not in combined:
        role = 'title'
    elif '副标题' in combined:
        role = 'subtitle'
    else:
        continue   # 不是标题/副标题形状，跳过

    # ── 获取文本框尺寸（XML 中单位是 EMU，1pt = 12700 EMU）──
    xfrm = sp.find(f'.//{A_XFRM}')
    box_w_emu = box_h_emu = None
    if xfrm is not None:
        ext = xfrm.find(A_EXT)
        if ext is not None:
            box_w_emu = int(ext.get('cx', 0))
            box_h_emu = int(ext.get('cy', 0))
    box_w_pt = (box_w_emu / 12700) if box_w_emu else 600
    box_h_pt = (box_h_emu / 12700) if box_h_emu else 100

    # ── 获取当前字号和字体风格 ──
    first_r = sp.find(f'.//{A_R}')
    current_sz = 24         # 模版无显式字号时的测量参考值
    has_explicit_sz = False  # 模版是否有显式字号
    font_path = FONT_TTF     # 默认常规体

    if first_r is not None:
        rpr = first_r.find(A_RPR)
        if rpr is not None:
            sz_str = rpr.get('sz')
            if sz_str:
                current_sz = int(sz_str) / 100      # XML 中存储的是百分之一磅
                has_explicit_sz = True
            if rpr.get('b') == '1':                 # 加粗
                font_path = FONT_TTF_BD

    shape_infos.append({
        'role': role,
        'box_w_pt': box_w_pt,
        'box_h_pt': box_h_pt,
        'current_sz': current_sz,
        'has_explicit_sz': has_explicit_sz,
        'font_path': font_path,
    })
    print(f"  模版识别: {role}  文本框={box_w_pt:.0f}×{box_h_pt:.0f}pt  原字号={current_sz}pt")


# ═══════════════════════════════════════════════════════════════
# 第三步：逐行生成 PPTX
# ═══════════════════════════════════════════════════════════════

OUT_DIR = '_生成结果'
os.makedirs(OUT_DIR, exist_ok=True)

for i, (folder, title, subtitle) in enumerate(rows):
    # ── 文件名：用文件夹名称命名，过滤掉非法字符 ──
    filename = f'{folder}.pptx'
    for ch in '\\/:*?"<>|':
        filename = filename.replace(ch, '_')

    # ── 复制模版 → 修改 slide1.xml → 写入新 PPTX ──
    out_path = os.path.join(OUT_DIR, filename)
    out_files = dict(template_files)         # 复制模版的所有文件
    slide_xml = copy.deepcopy(slide1_xml)    # 复制第一页 XML

    shape_idx = 0   # 指向 shape_infos，与迭代到的形状一一对应
    new_sz = 0
    lines = 0

    for sp in slide_xml.iter(P_SP):
        # 合并形状内所有文本，判断角色
        t_elems = list(sp.iter(A_T))
        combined = ''.join(t.text or '' for t in t_elems)

        if '标题' in combined and '副标题' not in combined:
            new_text = title
        elif '副标题' in combined:
            new_text = subtitle
        else:
            continue   # 不是目标形状，跳过

        # ── 清理多余的 run，只保留第一个（保留其字体格式）──
        r_elems = list(sp.iter(A_R))
        for r in r_elems[1:]:
            r.getparent().remove(r)

        # ── 写入新文本 ──
        first_t = r_elems[0].find(A_T)
        if first_t is not None:
            first_t.text = new_text

        # ── AutoFit 计算：根据文本长度和文本框尺寸算出最佳字号 ──
        info = shape_infos[shape_idx]
        new_sz, lines = calc_multiline(
            new_text, info['font_path'],
            info['box_w_pt'], info['box_h_pt'],
            info['current_sz'],
        )

        # ── 写入字号（仅模版有显式字号时才写入，否则继承模版主题）──
        rpr = r_elems[0].find(A_RPR)
        if rpr is not None and info['has_explicit_sz']:
            rpr.set('sz', str(int(new_sz * 100)))

        # ── 设置段落居中对齐 ──
        for p_elem in sp.iter(A_P):
            pPr = p_elem.find(A_PPR)
            if pPr is None:
                pPr = etree.SubElement(p_elem, A_PPR)
                p_elem.insert(0, pPr)
            pPr.set('algn', 'ctr')

        shape_idx += 1

    # ── 将修改后的 XML 写回文件字典 ──
    out_files['ppt/slides/slide1.xml'] = etree.tostring(
        slide_xml, xml_declaration=True, encoding='UTF-8', standalone=True)

    # ── 写出 PPTX 文件 ──
    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in out_files.items():
            zout.writestr(name, data)

    print(f"  [{i+1}/{len(rows)}] {filename}  (标题={new_sz}pt/{lines}行)")

print(f"\n[OK] 完成！共生成 {len(rows)} 个文件，保存在 {OUT_DIR}/ 目录")
